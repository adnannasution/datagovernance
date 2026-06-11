import openpyxl
import re

TAG_PATTERN = re.compile(r'\b[0-9][A-Z]{1,2}-[A-Z]{1,3}-[0-9]{3,4}[A-Z]?\b')

def detect_tag_numbers(text: str) -> list[str]:
    return list(set(TAG_PATTERN.findall(text.upper())))

def is_data_sheet(ws, sample_rows=10) -> tuple[bool, int]:
    """
    Deteksi apakah sheet ini sheet data tabular.
    Return (is_data, header_row_index)
    """
    non_empty_rows = 0
    max_cols_row = 0
    header_candidate = 0

    for i, row in enumerate(ws.iter_rows(max_row=sample_rows, values_only=True)):
        non_empty = sum(1 for c in row if c is not None)
        if non_empty > max_cols_row:
            max_cols_row = non_empty
            header_candidate = i

        if non_empty >= 2:
            non_empty_rows += 1

    # Kalau lebih dari setengah baris sampel terisi dan ada min 3 kolom → data sheet
    is_data = non_empty_rows >= (sample_rows * 0.4) and max_cols_row >= 3
    return is_data, header_candidate

def process_sheet_as_data(doc_id, ws, sheet_name, header_row_idx, chunk_index_start):
    """Proses sheet tabular → doc_table_rows + chunk per baris."""
    chunks = []
    table_rows_out = []
    chunk_index = chunk_index_start

    rows_iter = list(ws.iter_rows(values_only=True))
    if header_row_idx >= len(rows_iter):
        return chunks, table_rows_out, chunk_index

    headers = [str(h).strip() if h is not None else f"col{i}"
               for i, h in enumerate(rows_iter[header_row_idx])]

    for row_idx, row in enumerate(rows_iter[header_row_idx + 1:], start=1):
        if not any(c is not None for c in row):
            continue

        row_dict = {}
        parts = []
        for h, cell in zip(headers, row):
            if cell is not None and str(cell).strip():
                val = str(cell).strip()
                row_dict[h] = val
                parts.append(f"{h}: {val}")

        if not parts:
            continue

        row_text = " | ".join(parts)
        tags = detect_tag_numbers(row_text)
        tag_num = tags[0] if tags else None

        table_rows_out.append({
            "doc_id": doc_id,
            "halaman": None,
            "sheet_name": sheet_name,
            "row_index": row_idx,
            "row_data": row_dict,
            "tag_number": tag_num
        })

        chunks.append({
            "doc_id": doc_id,
            "chunk_index": chunk_index,
            "halaman": None,
            "slide_number": None,
            "slide_title": None,
            "sheet_name": sheet_name,
            "content": f"[Sheet: {sheet_name}] {row_text}",
            "embedding": None
        })
        chunk_index += 1

    return chunks, table_rows_out, chunk_index

def process_sheet_as_text(doc_id, ws, sheet_name, chunk_index_start):
    """Proses sheet narasi (cover, instruksi, dll) → chunk teks."""
    chunks = []
    chunk_index = chunk_index_start
    buffer = ""

    for row in ws.iter_rows(values_only=True):
        row_text = " | ".join(str(c).strip() for c in row if c is not None)
        if row_text.strip():
            buffer += row_text + "\n"
            if len(buffer) > 1000:
                chunks.append({
                    "doc_id": doc_id,
                    "chunk_index": chunk_index,
                    "halaman": None,
                    "slide_number": None,
                    "slide_title": None,
                    "sheet_name": sheet_name,
                    "content": f"[Sheet: {sheet_name}] {buffer.strip()}",
                    "embedding": None
                })
                chunk_index += 1
                buffer = ""

    if buffer.strip():
        chunks.append({
            "doc_id": doc_id,
            "chunk_index": chunk_index,
            "halaman": None,
            "slide_number": None,
            "slide_title": None,
            "sheet_name": sheet_name,
            "content": f"[Sheet: {sheet_name}] {buffer.strip()}",
            "embedding": None
        })
        chunk_index += 1

    return chunks, chunk_index

def process_excel(doc_id: int, file_path: str) -> dict:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    all_chunks = []
    all_table_rows = []
    chunk_index = 0
    sheet_meta = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        is_data, header_row_idx = is_data_sheet(ws)

        if is_data:
            chunks, table_rows, chunk_index = process_sheet_as_data(
                doc_id, ws, sheet_name, header_row_idx, chunk_index
            )
            sheet_meta.append({
                "name": sheet_name,
                "type": "data",
                "rows": len(table_rows)
            })
        else:
            chunks, chunk_index = process_sheet_as_text(
                doc_id, ws, sheet_name, chunk_index
            )
            table_rows = []
            sheet_meta.append({
                "name": sheet_name,
                "type": "text",
                "rows": len(chunks)
            })

        all_chunks.extend(chunks)
        all_table_rows.extend(table_rows)

    wb.close()

    all_tags = []
    for chunk in all_chunks:
        all_tags.extend(detect_tag_numbers(chunk["content"]))

    return {
        "chunks": all_chunks,
        "table_rows": all_table_rows,
        "auto_tags": list(set(all_tags)),
        "total_pages": len(wb.sheetnames),
        "processing_meta": {"sheets": sheet_meta}
    }
